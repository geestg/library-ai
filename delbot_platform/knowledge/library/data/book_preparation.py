from __future__ import annotations

import os
from typing import Any, Dict, List
import openpyxl


class BookPreparation:
    """
    Memuat dan mempersiapkan data katalog buku dari Excel.
    """
    def __init__(self, excel_path: str):
        self.excel_path = excel_path

    def load_books(self) -> List[Dict[str, Any]]:
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found at: {self.excel_path}")
            
        print(f"[PREPARE] Loading file: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Extract headers
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        
        books = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            title = row_dict.get("Title") or row_dict.get("judul")
            if not title or str(title).strip() == "":
                continue
            books.append(row_dict)
            
        print(f"[PREPARE] Loaded {len(books)} books successfully.")
        return books

    def validate_book(self, item: Dict[str, Any]) -> bool:
        """
        Validasi sederhana apakah item buku memiliki informasi dasar.
        """
        title = item.get("Title") or item.get("judul")
        return bool(title and str(title).strip() != "")
