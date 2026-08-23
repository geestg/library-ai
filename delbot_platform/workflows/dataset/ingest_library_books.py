"""
INGEST LIBRARY BOOKS
====================
Membaca dapus.xlsx dan mengupload semua buku ke Qdrant
collection 'library_books' dengan embedding nomic-embed-text (768 dim).

Jalankan dari dalam container:
  docker exec libraryai_backend python app/dataset/ingest_library_books.py
"""

import uuid
import sys

import openpyxl

from qdrant_client.models import PointStruct

from delbot_platform.research.retrieval.embedder import get_embedding
from delbot_platform.research.retrieval.qdrant_client import client, ensure_collection_exists
from delbot_platform.core.constants import LIBRARY_BOOKS_COLLECTION


# =========================================
# CONFIG
# =========================================

XLSX_PATH = "/app/app/dataset/dapus.xlsx"
BATCH_SIZE = 32


# =========================================
# LOAD EXCEL
# =========================================

def load_books(path: str):
    print(f"\n[LOAD] Reading: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    print(f"[LOAD] Columns: {headers}")

    books = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        title = row_dict.get("Title") or ""
        if not title.strip():
            continue
        books.append(row_dict)

    print(f"[LOAD] Total books loaded: {len(books)}")
    return books


# =========================================
# BUILD TEXT FOR EMBEDDING
# =========================================

def build_embed_text(book: dict) -> str:
    title = book.get("Title") or ""
    subject = book.get("Subject") or ""
    description = book.get("Deskripsi") or ""
    author = book.get("Author") or ""
    publisher = book.get("Publisher") or ""
    synopsis = book.get("Synopsis") or ""

    return f"""
Judul: {title}
Subjek: {subject}
Penulis: {author}
Penerbit: {publisher}
Deskripsi: {description}
Sinopsis: {synopsis}
""".strip()


# =========================================
# BUILD PAYLOAD
# =========================================

def build_payload(book: dict) -> dict:
    return {
        "title": book.get("Title") or "",
        "author": book.get("Author") or "",
        "publisher": book.get("Publisher") or "",
        "year": str(book.get("Published At") or ""),
        "subject": book.get("Subject") or "",
        "description": book.get("Deskripsi") or "",
        "synopsis": book.get("Synopsis") or "",
        "isbn": str(book.get("ISBN Number") or ""),
        "edition": str(book.get("Edition") or ""),
        "location": book.get("Location") or "",
        "language": book.get("Language") or "",
        "classification": book.get("Classification Number") or "",
        "program_study": book.get("Program Study") or "",
        "total_pages": int(book.get("Total Pages") or 0),
        "source_file": "dapus.xlsx",
        "text": build_embed_text(book),
    }


# =========================================
# MAIN INGEST
# =========================================

def ingest():
    print("\n" + "=" * 60)
    print("LIBRARY BOOKS INGEST")
    print("=" * 60)

    # Ensure collection exists dengan dim 768
    ensure_collection_exists(LIBRARY_BOOKS_COLLECTION, vector_size=768)

    # Load books
    books = load_books(XLSX_PATH)

    if not books:
        print("[ERROR] Tidak ada buku ditemukan di file Excel!")
        sys.exit(1)

    points = []
    failed = 0

    for idx, book in enumerate(books):
        title = book.get("Title", "")
        text = build_embed_text(book)

        if not text.strip():
            print(f"[SKIP] #{idx+1} - Teks kosong")
            continue

        try:
            embedding = get_embedding(text)
            payload = build_payload(book)

            point = PointStruct(
                id=str(uuid.uuid4()),
                vector=embedding,
                payload=payload,
            )

            points.append(point)

            if (idx + 1) % 50 == 0 or idx == 0:
                print(f"[EMBED] #{idx+1}/{len(books)} - {title[:60]}")

        except Exception as e:
            print(f"[ERROR] #{idx+1} '{title[:40]}': {e}")
            failed += 1
            continue

    print(f"\n[SUMMARY] Total embedded: {len(points)}, Failed: {failed}")

    if not points:
        print("[ERROR] Tidak ada points untuk di-upload!")
        sys.exit(1)

    # Upsert ke Qdrant dalam batch
    print(f"\n[UPLOAD] Uploading {len(points)} points to '{LIBRARY_BOOKS_COLLECTION}'...")

    for start in range(0, len(points), BATCH_SIZE):
        batch = points[start : start + BATCH_SIZE]
        end = min(start + BATCH_SIZE, len(points))

        client.upsert(
            collection_name=LIBRARY_BOOKS_COLLECTION,
            points=batch,
        )

        print(f"[UPSERT] Batch {start+1}-{end} ✓")

    print(f"\n[DONE] {len(points)} buku berhasil diindex ke '{LIBRARY_BOOKS_COLLECTION}'!")
    print("=" * 60)


# =========================================
# ENTRY POINT
# =========================================

if __name__ == "__main__":
    ingest()
